#! /usr/bin/env python3
"""This program will take an address as an input from the command line and
create an Excel file on the desktop with a specially formated table containing
comparable property sales in the area, complete with address, distance from
subject, price sold, date sold, price per SF, etc."""

#Import dependencies
import openpyxl
import os
import json
import requests
import datetime
import bs4
import http.client
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

#Change directory to desktop
os.chdir('/Users/spencercorwin/Desktop')

#Variable inputs
API_key = 'ONBOARD API HERE'
ZWSID = 'ZILLOW API HERE'
address1 = '2184+Canyon+Dr'
address2 = 'Costa+Mesa+CA'
zipCode = '92627'
radius = '0.06'
page = '1'
pagesize ='100'

#Use this to get all the properties in the area. Get their addresses and property IDs
def get_Onboard_properties_from_address(API_key, address1, address2, radius, page, pagesize):
    conn = http.client.HTTPSConnection("search.onboard-apis.com")
    headers = {'accept': "application/json",'apikey': API_key}
    conn.request("GET", "/propertyapi/v1.0.0/property/address?address1={}&address2={}&radius={}&page={}&pagesize={}".format(address1,address2,radius,page,pagesize), headers=headers) 
    res = conn.getresponse() 
    data = res.read()
    json_data = json.loads(data)
    properties = json_data['property']
    for i in range(len(properties)):
        allData[i] = {}
        allData[i]['Address'] = properties[i]['address']['line1'] #gives address line 1
        allData[i]['City, State'] = properties[i]['address']['line2'] #gives address line 2
        allData[i]['OnboardID'] = properties[i]['identifier']['obPropId'] #gives the Onboard property ID

#Use this to get sales history and property data for each address
def get_prop_data_sales_history_from_address(API_key, address1, address2):
    conn = http.client.HTTPSConnection("search.onboard-apis.com")
    headers = {'accept': "application/json",'apikey': API_key}
    conn.request("GET", "/propertyapi/v1.0.0/saleshistory/detail?address1={}&address2={}".format(address1, address2), headers=headers) 
    res = conn.getresponse() 
    data = res.read()
    json_data = json.loads(data)
    try:
        properties = json_data['property']
        propertyData['SF'] = properties[0]['building']['size']['bldgsize']    #gives building size
        propertyData['Beds'] = properties[0]['building']['rooms']['beds']   #gives the number of beds
        propertyData['Baths'] = properties[0]['building']['rooms']['bathstotal']    #gives the number of baths
        if len(properties[0]['salehistory']) == 1:
            propertyData['Sale Price'] = properties[0]['salehistory'][0]['amount']['saleamt']   #gives the most recent sales amount
            propertyData['Sale Date'] = properties[0]['salehistory'][0]['amount']['salerecdate']    #gives the most recent sale date
        elif len(properties[0]['salehistory']) > 1:
            propertyData['Sale Price'] = properties[0]['salehistory'][0]['amount']['saleamt']   #gives the most recent sales amount
            propertyData['Sale Date'] = properties[0]['salehistory'][0]['amount']['salerecdate']    #gives the most recent sale date
            propertyData['Prev Sale Price'] = properties[0]['salehistory'][1]['amount']['saleamt']  #gives the next most recent sales amount
            propertyData['Prev Sale Date'] = properties[0]['salehistory'][1]['amount']['salerecdate']   #gives the next most recent sales amount
    except:
        pass

#Use this to get a Zillow Property ID
def getSearchResults(address1, zipCode, ZWSID):
    url = 'http://www.zillow.com/webservice/GetSearchResults.htm?zws-id={}&address={}&citystatezip={}'.format(ZWSID,address1,zipCode)
    response = requests.get(url)
    response.raise_for_status()
    data = bs4.BeautifulSoup(response.text, 'lxml')
    elems = data.select('zpid')
    return (elems[0].getText())

#Get the Zestimate for a property with the Zillow Property ID
def getZestimate(ZWSID, ZPID):
    url = 'http://www.zillow.com/webservice/GetZestimate.htm?zws-id={}&zpid={}'.format(ZWSID,ZPID)
    response = requests.get(url)
    response.raise_for_status()
    data = bs4.BeautifulSoup(response.text, 'lxml')
    elems = data.select('zestimate amount')
    return (elems[0].getText())



#TODO: format excel sections to repond to the length of allData
#TODO: change variable address to address1, numOfComps
#TODO: make sure lengths of both Onboard functions are correct (may need to add 1 to each loop)
#TODO: remove Onboard property ID from first function?

#Create dictionary to store all data from above functions that will be used to fill the Excel doc
allData = {}

#Use first function to get the initial dictionary with all of the properties to be gathered
get_Onboard_properties_from_address(API_key, address1, address2, radius, page, pagesize)

#Use second function to get more info for each property into the dictionary allData
for i in range(len(allData)):
    propertyData = {}
    get_prop_data_sales_history_from_address(API_key, '+'.join(allData[i]['Address'].split(' ')), '+'.join(allData[i]['City, State'].split(' ')))
    try:
        allData[i]['SF'] = propertyData['SF']
        allData[i]['Beds'] = propertyData['Beds']
        allData[i]['Baths'] = propertyData['Baths']
        allData[i]['Sale Price'] = propertyData['Sale Price']
        allData[i]['Sale Date'] = propertyData['Sale Date']
        allData[i]['Prev Sale Price'] = propertyData['Prev Sale Price']
        allData[i]['Prev Sale Date'] = propertyData['Prev Sale Date']
    except:
        continue

#Create Excel workbook on Desktop
todaysDate = datetime.datetime.now()
todaysDateString = todaysDate.strftime('%m.%d.%y')      #get today's date in a nice format
headingFontObj = Font(bold=True, size=11, name='Calibri', underline='single')
wb = openpyxl.Workbook()                #create workbook on Desktop with formatted name
sheet = wb.active
headingsList = ['Address','Mi from Subject','Beds', 'Baths','SF','Sale Price','Price/SF','Sale Date','Zestimate','Zestimate/SF','Prev Sale Price','Prev Sale Date','Appreciation']

#Fill in the row with data from allData returned from getDeepComps
for i in range(len(allData)):
    try:
        sheet.cell(row=i+2,column=1).value = allData[i]['Address']
        sheet.cell(row=i+2,column=3).value = int(allData[i]['Beds'])
        sheet.cell(row=i+2,column=4).value = float(allData[i]['Baths'])
        sheet.cell(row=i+2,column=5).value = int(allData[i]['SF'])
        sheet.cell(row=i+2,column=6).value = int(allData[i]['Sale Price'])
        sheet.cell(row=i+2,column=8).value = allData[i]['Sale Date']
        sheet.cell(row=i+2,column=9).value = int(allData[i]['Zestimate'])
    except:
        continue

#Fill in the header row with the headings from headingsList
for c in range(1,len(headingsList)+1):
    sheet.cell(row=1,column=c).value = headingsList[c-1]
    sheet.cell(row=1,column=c).font = headingFontObj

#Create formulas within the table
for r in range(2,len(allData)+2):
    sheet.cell(row=r,column=7).value = '=F'+str(r)+'/E'+str(r)
    sheet.cell(row=r,column=10).value = '=I'+str(r)+'/E'+str(r)
    sheet.cell(row=r,column=13).value = '=(F'+str(r)+'-K'+str(r)+')/K'+str(r)

#Create averages and totals at the bottom of the table
bottomFontObj = Font(bold=True, size=11, name='Calibri')
averagedColumns = [5,6,7,9,10]
avgRow = len(allData)+2
sheet.cell(row=avgRow,column=4).value = 'Averages:'
sheet.cell(row=avgRow,column=4).font = bottomFontObj
for c in averagedColumns:
    sheet.cell(row=avgRow,column=c).value = '=average({0}2:{0}{1})'.format(get_column_letter(c), len(allData)+1)
    sheet.cell(row=avgRow,column=c).font = bottomFontObj

#Create borders around all cells except the last row
borderObj = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
for c in range(1,sheet.max_column+1):
    for r in range(1,sheet.max_row):
        sheet.cell(row=r,column=c).border = borderObj

#Fill cells with white color
whiteFill = PatternFill('solid', fgColor='00FFFFFF')
for c in range(1,sheet.max_column+1):
    for r in range(1,sheet.max_row+1):
        sheet.cell(row=r,column=c).fill = whiteFill

#Resize columns
for c in range(1,sheet.max_column):
    max_length = len(str(sheet.cell(row=1,column=c).value))
    for r in range(1,sheet.max_row):
        try:
            if len(str(sheet.cell(row=r,column=c).value)) > max_length:
                max_length = len(str(sheet.cell(row=r,column=c).value))
        except:
            continue
    sheet.column_dimensions[get_column_letter(c)].width = max_length
    
#Save and close the file with a formatted name
wb.save(' '.join(address1.split('+'))+' Comps '+todaysDateString+'.xlsx')
print('I\'ve done your bidding, human.')
